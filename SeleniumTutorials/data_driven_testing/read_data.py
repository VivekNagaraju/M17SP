import openpyxl


filepath = "C:\\Users\\admin\\Desktop\\OrangeHRM_Login_DDT.xlsx"
my_workbook = openpyxl.load_workbook(filepath)
# my_active_sheet = my_workbook.active # Get the currently active sheet
my_active_sheet = my_workbook["Sheet1"] # Get the sheet by name
total_rows = my_active_sheet.max_row
print(total_rows)

total_columns = my_active_sheet.max_column
print(total_columns)

username = my_active_sheet.cell(2, 1).value
password = my_active_sheet.cell(2, 2).value
print(username, password)

username = my_active_sheet.cell(3, 1).value
password = my_active_sheet.cell(3, 2).value
print(username, password)

username = my_active_sheet.cell(4, 1).value
password = my_active_sheet.cell(4, 2).value
print(username, password)

username = my_active_sheet.cell(5, 1).value
password = my_active_sheet.cell(5, 2).value
print(username, password)
